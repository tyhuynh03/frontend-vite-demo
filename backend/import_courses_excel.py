#!/usr/bin/env python
"""
Script để import dữ liệu khóa học từ file Excel vào database
Chạy: python import_courses_excel.py
"""

import os
import sys
import django
import openpyxl

# Setup Django environment
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

from core.models import Course

# Level giờ đây có thể chứa bất kỳ giá trị nào từ Excel

def import_courses_from_excel(excel_file=None):
    """Import courses from Excel file"""
    
    if excel_file is None:
        excel_file = os.path.join(os.path.dirname(__file__), 'Khoa_hoc.xlsx')
    
    if not os.path.exists(excel_file):
        print(f"❌ Không tìm thấy file {excel_file}")
        return
    
    try:
        # Đọc file Excel
        workbook = openpyxl.load_workbook(excel_file)
        worksheet = workbook.active
        
        print(f"📖 Đang đọc file Excel: {excel_file}")
        
        # Lấy header từ dòng đầu tiên
        headers = []
        for col in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=1, column=col).value
            if cell_value:
                headers.append(cell_value.strip().lower())
        
        print(f"📋 Các cột được tìm thấy: {headers}")
        
        # Kiểm tra các cột bắt buộc
        required_columns = ['title', 'level', 'category', 'description', 'content']
        missing_columns = [col for col in required_columns if col not in headers]
        
        if missing_columns:
            print(f"❌ Thiếu các cột bắt buộc: {missing_columns}")
            print(f"📋 Các cột cần có: {required_columns}")
            return
        
        # Đọc dữ liệu từ dòng thứ 2 trở đi
        courses_imported = 0
        courses_skipped = 0
        
        for row in range(2, worksheet.max_row + 1):
            # Kiểm tra dòng có dữ liệu không
            title_cell = worksheet.cell(row=row, column=1)
            if not title_cell.value:
                continue
            
            # Tạo dictionary cho course
            course_data = {}
            for col, header in enumerate(headers, 1):
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value:
                    course_data[header] = str(cell_value).strip()
            
            # Kiểm tra dữ liệu bắt buộc
            if not all(course_data.get(col) for col in required_columns):
                print(f"⚠️  Dòng {row}: Thiếu dữ liệu bắt buộc, bỏ qua")
                courses_skipped += 1
                continue
            
            # Kiểm tra xem khóa học đã tồn tại chưa
            existing_course = Course.objects.filter(title=course_data['title']).first()
            
            if existing_course:
                print(f"⚠️  Khóa học '{course_data['title']}' đã tồn tại, bỏ qua...")
                courses_skipped += 1
                continue
            
            # Level giờ đây có thể chứa bất kỳ giá trị nào từ Excel
            
            # Tạo khóa học mới
            try:
                course = Course.objects.create(
                    title=course_data['title'],
                    level=course_data['level'],
                    category=course_data['category'],
                    description=course_data['description'],
                    content=course_data['content'],
                    is_active=True
                )
                print(f"✅ Đã tạo khóa học: {course.title}")
                courses_imported += 1
                
            except Exception as e:
                print(f"❌ Lỗi khi tạo khóa học '{course_data['title']}': {e}")
                courses_skipped += 1
        
        print(f"\n📊 Kết quả import:")
        print(f"   ✅ Đã import: {courses_imported} khóa học")
        print(f"   ⚠️  Đã bỏ qua: {courses_skipped} khóa học")
        print(f"   📚 Tổng cộng trong database: {Course.objects.count()} khóa học")
        
    except Exception as e:
        print(f"❌ Lỗi khi đọc file Excel: {e}")

def show_help():
    """Hiển thị hướng dẫn sử dụng"""
    print("🚀 HƯỚNG DẪN SỬ DỤNG IMPORT EXCEL")
    print("=" * 50)
    print("1. Tạo file Excel mẫu:")
    print("   python create_excel_template.py")
    print("")
    print("2. Chỉnh sửa file courses_sample.xlsx")
    print("   - Thêm/sửa/xóa khóa học")
    print("   - Đảm bảo đủ các cột: title, level, category, description, content")
    print("")
    print("3. Import vào database:")
    print("   python import_courses_excel.py")
    print("")
    print("📋 CÁC LEVEL HỢP LỆ:")
    print("   - tieu-hoc (Tiểu học)")
    print("   - thcs (THCS)")
    print("   - thpt (THPT)")
    print("   - dai-hoc (Đại học)")
    print("   - cao-hoc (Cao học)")
    print("   - giao-vien (Giáo viên)")
    print("   - can-bo (Cán bộ)")
    print("   - moi-tuoi (Mọi lứa tuổi)")

if __name__ == '__main__':
    if len(sys.argv) > 1 and sys.argv[1] in ['-h', '--help', 'help']:
        show_help()
    else:
        excel_file = sys.argv[1] if len(sys.argv) > 1 else None
        import_courses_from_excel(excel_file)
